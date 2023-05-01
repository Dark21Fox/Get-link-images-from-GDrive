import os
import sys
import magic
import openpyxl
import traceback
from design.form import Ui_Widget
from PySide6 import QtGui, QtCore
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from PySide6.QtWidgets import QApplication, QWidget, QMessageBox, QTableWidgetItem, QFileDialog, QHeaderView, QListWidgetItem

SCOPES = ['https://www.googleapis.com/auth/drive']


class MainWindow(QWidget):
    def __init__(self):
        self.acc_file = None
        self.drive_service = None
        super(MainWindow, self).__init__()
        self.ui = Ui_Widget()
        self.ui.setupUi(self)
        self.ui.toolButton.clicked.connect(self.find_json)
        self.ui.pushButton.clicked.connect(self.add_image_drive)
        self.ui.pushButton_txt.clicked.connect(self.export_txt)
        self.ui.pushButton_excel.clicked.connect(self.export_excel)

    def find_json(self):
        file, _ = QFileDialog.getOpenFileName(self, 'Open File', './', "Json (*.json)")
        self.ui.lineEdit.setText(file)
        self.acc_file = self.ui.lineEdit.text()
        credentials = service_account.Credentials.from_service_account_file(self.acc_file, scopes=SCOPES)
        self.drive_service = build('drive', 'v3', credentials=credentials)
        if self.ui.lineEdit.text():
            self.ui.widget.setEnabled(1)
            self.ui.frame.setEnabled(1)
            self.folder_in_acc()
        else:
            QMessageBox.critical(self, "Ошибка ", "Не выбран JSON файл", QMessageBox.Ok)

    def export_txt(self):
        file = open("Google_Drive.txt", "w")
        file.write(self.ui.textBrowser_url.toPlainText())
        file.close()

    def export_excel(self):
        file = 'Google_Drive.xlsx'
        wb = openpyxl.Workbook()
        wb.save(file)
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[0]
        urls = self.ui.textBrowser_url.toPlainText().split("\n")
        name = self.ui.textBrowser_name.toPlainText().split("\n")
        self.add_to_excel(ws, name, 1)
        self.add_to_excel(ws, urls, 2)
        wb.save(file)
        QMessageBox.information(self, "Успех", "Успешно создан xlsx", QMessageBox.Ok)

    def add_to_excel(self, ws, data, column):
        for r, item in enumerate(data, start=1):
            ws.cell(row=r, column=column).value = item

    def folder_in_acc(self):
        results = self.drive_service.files().list(pageSize=1000,
                                                  fields="files(id, name, mimeType)",
                                                  q="mimeType = 'application/vnd.google-apps.folder'").execute()
        res = results.get('files')
        for i, item in enumerate(res, start=0):
            item_lw = QListWidgetItem(f"{item.get('name')} : {item.get('id')}")
            self.ui.listWidget.addItem(item_lw)

    def add_image_drive(self):
        try:
            try:
                folder = [self.ui.listWidget.currentItem().text().split(" : ")[1]]
            except:
                QMessageBox.critical(self, "Ошибка", "Выберите директорию!", QMessageBox.Ok)
            self.ui.pushButton_excel.setEnabled(1)
            self.ui.pushButton_txt.setEnabled(1)
            file, _ = QFileDialog.getOpenFileNames(self, 'Open File', './', "Image (*.png *.jpg *jpeg)")
            self.ui.tableWidget.setColumnCount(1)
            self.ui.tableWidget.setRowCount(len(file))
            self.ui.progressBar.setValue(10)
            self.ui.tableWidget.setIconSize(QtCore.QSize(150, 200))
            self.ui.label_5.setText(f"Выбрано картинок : {len(file)}")
            self.ui.progressBar.setValue(30)
            column = 0
            for image_path in file:
                item = QTableWidgetItem()
                self.ui.tableWidget.setItem(0, column, item)
                item.setIcon(QtGui.QIcon(image_path))
                column += 1
            self.ui.progressBar.setValue(40)
            self.ui.tableWidget.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.ui.tableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            self.ui.tableWidget.horizontalHeader().setMinimumSectionSize(0)
            self.ui.progressBar.setValue(60)
            for image in file:
                basename = os.path.basename(image)
                mime = magic.from_file(image, mime=True)
                file_metadata = {
                    "name": basename,
                    "mimeType": mime,
                    "parents": folder}
                media = MediaFileUpload(image, mimetype=mime, resumable=True)
                created = self.drive_service.files().create(body=file_metadata, media_body=media,
                                                            fields="id,webViewLink").execute()
                file_permission = {"role": "reader", "type": "anyone"}
                self.drive_service.permissions().create(body=file_permission, fileId=created.get("id")).execute()
                link = created.get("webViewLink")
                full = 'https://drive.google.com/uc?export=view&id=' + link[link.find("d/") + 2:link.rfind("/")]
                self.ui.textBrowser_name.append(basename)
                self.ui.textBrowser_url.append(full)
                print(full)
            QMessageBox.information(self, "Успешно ", "Загрузка фото выполнена", QMessageBox.Ok)
            self.ui.progressBar.setValue(100)
            return file
        except (UnicodeDecodeError, OSError) as exe:
            print(f'Произошла ошибка: {exe}', traceback.format_exc())
            QMessageBox.critical(self, "Ошибка ",
                                 "Путь к файлам НЕ должен содержать буквы кириллицы!\n Или \nнеподдерживаемый формат файла",
                                 QMessageBox.Ok)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = MainWindow()
    widget.show()
    sys.exit(app.exec())
